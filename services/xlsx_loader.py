from abstract_file_loader import AbstractFileLoader
import os
from openpyxl import load_workbook

class XLSXLoader(AbstractFileLoader):
    
    def __init__(self):
        super().__init__()
    
    def load(self, file_path, file_name):
        """
        Load the content from the specified XLSX file path.
        :param file_path: The path to the XLSX file to check.
        :param file_name: The name of the XLSX file to check.
        :return: A dictionary containing the file's metadata and content.
        """
        full_path = os.path.join(file_path, file_name)
        workbook = load_workbook(full_path, data_only=True)
        
        content = ""
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            for row in worksheet.iter_rows(values_only=True):
                row_content = " ".join(str(cell) for cell in row if cell is not None)
                content = content + row_content
        
        doc = {
            "name": file_name,
            "path": full_path,
            "extension": "xlsx",
            "content": content.strip().replace("\n", " "),
            "size_bytes": os.path.getsize(full_path)
        }
        
        return doc
    
    def exists(self, file_path, file_name):
        """
        Check if the specified XLSX file exists.

        :param file_path: The path to the XLSX file to check.
        :param file_name: The name of the XLSX file to check.
        :return: True if the file exists, False otherwise.
        """
        return os.path.exists(os.path.join(file_path, file_name)) and os.path.isfile(os.path.join(file_path, file_name))